import streamlit as st
import pandas as pd
import mysql.connector
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import date
import plotly.express as px

# --- DB Configuration ---
DB_CONFIG = {
    'host': '172.16.128.79',
    'user': 'usr4mis',
    'password': 'usr4MIS#@!',
    'database': 'volare',
    'port': 3307
}

# --- Page Title ---
st.title("📊 PTP and Agent Posted Payments Dashboard - BPI CARDS XDAYS")

# --- Date Range Picker ---
col1, col2 = st.columns(2)
with col1:
    start_date = st.date_input("Start Date", value=date(2025, 4, 10))
with col2:
    end_date = st.date_input("End Date", value=date(2025, 4, 14))

if start_date > end_date:
    st.error("🚫 Start date must be before end date.")
    st.stop()

# --- DB Connection Helper ---
def create_connection():
    return mysql.connector.connect(**DB_CONFIG)

# --- Fetch PTP Data ---
@st.cache_data(ttl=600, show_spinner="Fetching PTP records...")
def fetch_ptp_data(start_date, end_date):
    query = """
        SELECT DISTINCT
            CONCAT('Cycle ', RIGHT(debtor.cycle, 2)) AS `CYCLE`,
            debtor.card_no AS `CH CODE`,
            debtor.account AS `ACCOUNT NUMBER`,
            debtor.name AS `NAME`,
            followup.remark_by AS `AGENT CODE`,
            followup.status_code AS `STATUS CODE`,
            followup.remark AS `REMARKS`,
            debtor.ptp_amount AS `PTP AMOUNT`,
            debtor.ptp_date AS `PTP DATE`,
            debtor.balance AS `OB`,
            followup.datetime AS `DISPO DATE`,
            debtor.placement AS `FINONE ID`,
            debtor.is_locked AS `IS LOCKED`,
            debtor.is_aborted AS `IS ABORTED`
        FROM debtor
        LEFT JOIN debtor_followup ON debtor_followup.debtor_id = debtor.id
        LEFT JOIN followup ON followup.id = debtor_followup.followup_id
        LEFT JOIN `user` ON `user`.id = followup.remark_by_id
        WHERE debtor.client_name LIKE '%BPI CARDS XDAYS%'
          AND followup.status_code LIKE 'PTP%'
          AND DATE(followup.`date`) BETWEEN %s AND %s
          AND followup.remark_by = 'GTCUSTODIO'
        ORDER BY followup.datetime DESC
    """
    conn = create_connection()
    try:
        return pd.read_sql(query, conn, params=(start_date, end_date))
    finally:
        conn.close()

# --- Fetch CURED/Posted Payments Data ---
@st.cache_data(ttl=600, show_spinner="Fetching Agent Posted Payments...")
def fetch_cured_data(start_date, end_date):
    query = """
        SELECT DISTINCT
            CONCAT('Cycle ', RIGHT(debtor.cycle, 2)) AS `CYCLE`,
            debtor.card_no AS `CH CODE`,
            debtor.account AS `ACCOUNT NUMBER`,
            followup.remark AS `REMARKS`,
            followup.remark_by AS `AGENT CODE`,
            followup.status_code AS `STATUS CODE`,
            debtor.ptp_amount AS `PTP AMOUNT`,
            debtor.ptp_date AS `PTP DATE`,
            debtor.balance AS `OB`,
            followup.datetime AS `DISPO DATE`,
            debtor.placement AS `FINONE ID`,
            debtor.is_locked AS `IS LOCKED`,
            debtor.is_aborted AS `IS ABORTED`
        FROM debtor
        LEFT JOIN debtor_followup ON debtor_followup.debtor_id = debtor.id
        LEFT JOIN followup ON followup.id = debtor_followup.followup_id
        LEFT JOIN `user` ON `user`.id = followup.remark_by_id
        WHERE debtor.client_name LIKE '%BPI CARDS XDAYS%'
          AND followup.status_code IN (
              'PAYMENT - UNPOSTED_MISPOSTED PAYMENTS',
              'PAYMENT - CURED',
              'PAYMENT - INSUFFIECIENT PAYMENT'
          )
          AND DATE(followup.`date`) BETWEEN %s AND %s
          AND followup.remark_by = 'GTCUSTODIO'
          AND followup.remark NOT LIKE '%MSPM%'
        ORDER BY followup.datetime DESC
    """
    conn = create_connection()
    try:
        return pd.read_sql(query, conn, params=(start_date, end_date))
    finally:
        conn.close()

# --- Fetch Data ---
ptp_df = fetch_ptp_data(start_date, end_date)
cured_df = fetch_cured_data(start_date, end_date)

# --- Display PTP Records ---
st.header("📋 PTP Records")
if ptp_df.empty:
    st.warning("No PTP records found for the selected date range.")
else:
    ptp_df["ACCOUNT NUMBER"] = ptp_df["ACCOUNT NUMBER"].astype(str).str.zfill(10)
    st.dataframe(ptp_df, use_container_width=True)

    st.subheader("📈 PTP Status Count by Agent")
    ptp_status_count = ptp_df.groupby(['AGENT CODE', 'STATUS CODE']).size().reset_index(name='Count')
    fig_ptp = px.bar(
        ptp_status_count,
        x='AGENT CODE',
        y='Count',
        color='STATUS CODE',
        title="PTP Status Count by Agent",
        labels={'Count': 'Number of Records'}
    )
    fig_ptp.update_layout(xaxis_tickangle=-45)
    st.plotly_chart(fig_ptp)

# --- Display Agent Posted Payments (CURED) Records ---
st.header("🧾 Agent Posted Payments (CURED) Records")
if cured_df.empty:
    st.warning("No Agent Posted Payments (CURED) records found for the selected date range.")
else:
    cured_df["ACCOUNT NUMBER"] = cured_df["ACCOUNT NUMBER"].astype(str).str.zfill(10)
    st.dataframe(cured_df, use_container_width=True)

# --- Download Button (Excel with Two Sheets) ---
if not ptp_df.empty or not cured_df.empty:
    output = BytesIO()
    wb = Workbook()

    # PTP Sheet
    ws_ptp = wb.active
    ws_ptp.title = "PTP Records"
    if not ptp_df.empty:
        for r in dataframe_to_rows(ptp_df, index=False, header=True):
            ws_ptp.append(r)
        for col in ws_ptp.columns:
            max_len = max(len(str(cell.value)) for cell in col if cell.value)
            ws_ptp.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
    else:
        ws_ptp.append(["No data found for PTP records."])

    # CURED Sheet
    ws_cured = wb.create_sheet("Agent Posted Payments")
    if not cured_df.empty:
        for r in dataframe_to_rows(cured_df, index=False, header=True):
            ws_cured.append(r)
        for col in ws_cured.columns:
            max_len = max(len(str(cell.value)) for cell in col if cell.value)
            ws_cured.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
    else:
        ws_cured.append(["No data found for Agent Posted Payments."])

    wb.save(output)
    st.download_button(
        label="⬇️ Download Excel (PTP + Agent Posted Payments)",
        data=output.getvalue(),
        file_name="PTP_and_Agent_Posted_Payments.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
